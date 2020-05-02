from handlers.common import BaseHandler, authenticated_async
from tornado import gen
from tornado.web import app_log as log
from handlers.common import AddLogs
from openpyxl import load_workbook
from settings import online_mail_to
import time
import os
import json
import uuid


"""
公共接口
"""


# 公共信息
class CommonInfoHandler(BaseHandler):
    @gen.coroutine
    def get(self):
        AddLogs().add_logs(ip=self.request.remote_ip)
        common_info = yield self.option.get_options_list(o_type='common', status=1)
        common = dict(company='鸥鹏斯塔克', sysName='斯塔克综合测试管理平台',
                      sysDesc='让测试更高效 使测试更简单', emailExt='')
        auto_login = False
        for row in common_info:
            if row.name == 'company':
                common['company'] = row.value
            if row.name == 'sysName':
                common['sysName'] = row.value
            if row.name == 'sysDesc':
                common['sysDesc'] = row.value
            if row.name == 'emailExt' and row.value:
                common['emailExt'] = '@{}'.format(row.value)
            if row.name == 'autoLogin':
                auto_login = True if row.value == 'yes' else False
        msg = dict(status='SUCCESS', message='', data=dict(
            company=common['company'], sysName=common['sysName'],
            sysDesc=common['sysDesc'], emailExt=common['emailExt'], autoLogin=auto_login))
        self.write_json(msg)


# 团队列表
class DepartmentsHandler(BaseHandler):
    @gen.coroutine
    def get(self):
        teams = yield self.option.get_options_list(o_type='teams', status=1)
        groups = []
        for row in teams:
            if row.name == 'company':
                company = json.loads(row.value)
                company = dict(value=row.id, label=company['name'], children=[])
                groups.append(company)
                teams.remove(row)
        for x in groups:
            company = x['value']
            for y in teams:
                if y.name == 'department':
                    department = json.loads(y.value)
                    if department['up'] == company:
                        department = dict(value=y.id, label=department['name'], children=[])
                        x['children'].append(department)
                        teams.remove(y)
        for x in groups:
            for y in x['children']:
                department = y['value']
                for z in teams:
                    if z.name == 'team':
                        team = json.loads(z.value)
                        if team['up'] == department:
                            team = dict(value=z.id, label=team['name'])
                            y['children'].append(team)
        msg = dict(status='SUCCESS', message='', data=groups or [dict(value=0, label='鸥鹏斯塔克')])
        self.write_json(msg)


# 文件上传接口
class UploadFiles(BaseHandler):
    @authenticated_async
    @gen.coroutine
    def get(self, op=None, name=None):
        AddLogs().add_logs(ip=self.request.remote_ip)
        if op == 'media' and name == 'files':
            media_path = os.path.join(self.settings.get('static_path'), op, name, str(self.current_user.id))
            data = list()
            if os.path.isdir(media_path):
                files_list = os.listdir(media_path)
                for file in files_list:
                    file_ext = file.split('.', maxsplit=1)
                    uri = self.static_url('{}/{}/{}/{}'.format(op, name, str(self.current_user.id), file))
                    if file_ext[-1].lower() == 'mp3':
                        data.append(dict(id=file_ext[0], type='AUDIO', url=uri))
                    elif file_ext[-1].lower() == 'mp4':
                        data.append(dict(id=file_ext[0], type='VIDEO', url=uri))
                    else:
                        data.append(dict(id=file_ext[0], type='IMAGE', url=uri))
            self.write_json(dict(status='SUCCESS', message='', data=data))
        else:
            self.write_json(dict(status='FAIL', message='操作类型错误', data=''))

    @authenticated_async
    @gen.coroutine
    def post(self, op=None, name=None):
        AddLogs().add_logs(ip=self.request.remote_ip, op_type='active')
        if not op and not name:
            return self.write_json(dict(status='FAIL', message='参数不正确', data=''))
        if op == 'media' and name == 'delete':
            data = yield self.get_request_body_to_json()
            media_path = os.path.join(self.settings.get('static_path'), op, 'files', str(self.current_user.id))
            if os.path.isdir(media_path):
                files_list = os.listdir(media_path)
                for file in files_list:
                    if file.split('.', maxsplit=1)[0] in data.get('keys'):
                        log.info('删除媒体文件#{}'.format(os.path.join(media_path, file)))
                        os.remove(os.path.join(media_path, file))
            return self.write_json(dict(status='SUCCESS', message='', data=op))
        files = self.request.files
        if op == 'media' and name == 'local':
            return self.write_json(dict(state="SUCCESS",
                                        list=[{'source': url, 'url': url, 'state': "SUCCESS"} for url in self.get_arguments('source[]')]))
        flag = False
        msg = '上传失败!'
        uri = ''
        if files:
            if op == 'images':
                images_path = os.path.join(self.settings.get('static_path'), op)
                images = files.get(name)
                if name == 'avatar' and images:
                    images = images[0]
                    filename = images.filename
                    ext = filename.split('.')[-1]
                    filename = '{}.{}'.format(int(time.time()*10000000), ext)
                    file_path = os.path.join(images_path, name)
                    flag = yield self.__save_file(file_path, filename, images.body)
                    if flag:
                        msg = '上传成功!'
                        log.info('文件上传成功, 保存在 {}'.format(flag))
                        uri = self.static_url('{}/{}/{}'.format(op, name, filename))
                        profile = json.loads(self.current_user.profile)
                        profile['avatar'] = uri
                        self.user.edit_user(uid=self.current_user.id, profile=profile)
                else:
                    flag = False
            elif op == 'files':
                files_path = os.path.join(self.settings.get('static_path'), op)
                files_list = files.get(name)
                test_type = self.get_argument('test_type', 'caseG')
                if name == 'testCase' and files_list:
                    file = files_list[0]
                    filename = file.filename
                    ext = filename.split('.')[-1]
                    filename = '{}.{}'.format(int(time.time()*10000000), ext)
                    file_path = os.path.join(files_path, name)
                    flag = file_path = yield self.__save_file(file_path, filename, file.body)
                    if flag:
                        log.info('文件上传成功, 保存在 {}'.format(file_path))
                        if os.path.isfile(file_path):
                            workbook = load_workbook(file_path, read_only=True)
                            try:
                                table = workbook[workbook.sheetnames[0]]
                                if table.max_row > 1 and table.max_column == 11:
                                    count_add = 0
                                    count_edit = 0
                                    for i in range(1, table.max_row):
                                        cid = table['B{}'.format(i+1)].value
                                        project = table['D{}'.format(i+1)].value
                                        module = table['E{}'.format(i+1)].value
                                        title = table['F{}'.format(i+1)].value
                                        description = table['G{}'.format(i+1)].value
                                        expected = table['H{}'.format(i+1)].value
                                        status = table['I{}'.format(i+1)].value
                                        author = table['J{}'.format(i+1)].value
                                        executor = table['K{}'.format(i+1)].value
                                        desc = dict(description=description or '', expected=expected or '',
                                                    function='' or '', executor=executor or '',
                                                    title=title or '', module=module or '', userId=self.current_user.id,
                                                    author=author or self.current_user.realname or '')
                                        if test_type == 'case':
                                            status = 2 if status == '已转自动化' else 1 if status == '手动执行' else 0
                                        else:
                                            status = 2 if status == '已实现' else 1 if status == '开发中' else 0
                                        if not cid:
                                            key, msg = yield self.setting.add_setting(
                                                s_type=test_type, name=cid or '', value=desc, status=status,
                                                project=project and project.strip())
                                            if key:
                                                count_add += 1
                                        else:
                                            projects, total = yield self.setting.get_settings_list(
                                                name=cid, project=project, limit=None, s_type=test_type)
                                            if projects:
                                                for p in projects:
                                                    desc['function'] = json.loads(p.value).get('function') or ''
                                                    desc['userId'] = json.loads(p.value).get('userId') or ''
                                                    desc['urls'] = json.loads(p.value).get('urls') or list()
                                                    key, msg = yield self.setting.edit_setting(
                                                        sid=p.id, value=desc, status=status)
                                                    if key:
                                                        count_edit += 1
                                            else:
                                                key, msg = yield self.setting.add_setting(
                                                    s_type=test_type, name=cid, value=desc, project=project, status=status)
                                                if key:
                                                    count_add += 1
                                    msg = '测试用例成功导入 {} 条, 编辑 {} 条!'.format(count_add, count_edit)
                                else:
                                    flag = False
                                    msg = '没有用例数据或用例模板格式不对, 请检查后重新导入!'
                            except Exception as e:
                                log.error(e)
                                flag = False
                                msg = '{}#{}'.format('测试用例导入失败', e)
                            workbook.close()
                            try:
                                os.remove(file_path)
                            except os.error as e:
                                log.info(e)
                        else:
                            flag = False
                            msg = '文件上传成功但保存失败!'
                else:
                    flag = False
            else:
                media_path = os.path.join(self.settings.get('static_path'), op)
                file = files.get(name)
                if name == 'files' or name == 'file' and file:
                    name = 'files'
                    file = file[0]
                    filename = file.filename
                    ext = filename.split('.')[-1]
                    filename = '{}.{}'.format(str(uuid.uuid1()), ext)
                    file_path = os.path.join(media_path, name, str(self.current_user.id))
                    flag = yield self.__save_file(file_path, filename, file.body)
                    if flag:
                        msg = '上传成功!'
                        log.info('文件上传成功, 保存在 {}'.format(flag))
                        uri = self.static_url('{}/{}/{}/{}'.format(op, name, str(self.current_user.id), filename))
                else:
                    flag = False
        if flag:
            self.write_json(dict(status='SUCCESS', message=msg, data=uri,
                                 state='SUCCESS', url=uri, title=self.get_body_argument('name', ''), original=self.get_body_argument('name', '')))
        else:
            self.write_json(dict(status='FAIL', message=msg, data='', state=msg, url='', title='', original=''))

    @gen.coroutine
    def __save_file(self, file_path, filename, data):
        try:
            if not os.path.isdir(file_path):
                os.makedirs(file_path)
            file_path = os.path.join(file_path, filename)
            with open(file_path, 'wb') as fp:
                fp.write(data)
            return file_path
        except Exception as e:
            log.error(e)
            return None


# 邮件发送接口
class SendMailHandler(BaseHandler):
    @authenticated_async
    @gen.coroutine
    def post(self, op=None):
        AddLogs().add_logs(ip=self.request.remote_ip, op_type='active')
        if op == 'online':
            data = yield self.get_request_body_to_json()
            content = data.get('content')
            if content:
                html = '''
                <p>Dear All:</p>
                <p>&nbsp;&nbsp;以下是近期的线上问题汇报，请查收~</p>
                <p>&nbsp;&nbsp;本信息由测试部自动化汇集工具自动创建于{}</p>
                <p>{}</p>
                <p>&nbsp;&nbsp;如您需要查看更多信息，请访问综合测试管理平台。</p>
                '''.format(time.strftime('%Y-%m-%d %H:%M:%S'), content)
                res, msg = yield self.common_func.send_email(
                    subject='测试部门线上问题汇报_{}'.format(time.strftime('%Y-%m-%d %H:%M:%S')), content=html, to=online_mail_to)
            else:
                res, msg = False, '邮件内容不能为空!'
            if res:
                msg = dict(status='SUCCESS', message='邮件发送成功!', data='')
            else:
                log.info(msg)
                msg = dict(status='FAIL', message=str(msg), data='')
        else:
            msg = dict(status='FAIL', message='操作类型错误!', data='')
        self.write_json(msg)
